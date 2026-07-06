"""One-time importer: load current positions + completed trades from the
Google-Sheets export into Postgres.

Usage:
    python seed_from_sheet.py ["C:\\path\\to\\Christian - Stock Trading for 2026.xlsx"]

Idempotent: clears lot/ticker/completed_trade and re-imports. The app owns the
data after this; re-run only if you want to re-seed from a fresh sheet export.

Sheet layout (decoded):
  Stock Data!E8:E208  -> ticker symbols (one per Longs block)
  Longs               -> 16-row block per ticker, block i starts at row 6+16*i;
                         lot rows are block_start+3 .. block_start+12 with
                         D = "MM/DD/YYYY TICKER", E = shares, F = buy price
  Long Log rows 7+    -> completed trades: B date+ticker, C shares, D buy,
                         F sell, H completed date, I ticker
"""
from __future__ import annotations

import asyncio
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.db import SessionLocal, init_db
from app.db.models import CompletedTrade, Lot, Ticker
from app.schwab.auth import load_client


def resolve_llc_hash() -> str | None:
    """The LLC is the CASH account (the personal trading account is MARGIN)."""
    client = load_client()
    if client is None:
        return None
    for n in client.get_account_numbers().json():
        h = n["hashValue"]
        try:
            sa = client.get_account(h).json().get("securitiesAccount", {})
            if sa.get("type") == "CASH":
                return h
        except Exception:
            continue
    return None

DEFAULT_XLSX = Path(
    r"C:\Users\dodar\Downloads\Christian - Stock Trading for 2026.xlsx"
)


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    token = str(value).strip().split(" ")[0]  # "03/18/2026 RCAT" -> "03/18/2026"
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def read_positions(path: Path):
    """Return {symbol: [(rung, buy_date, shares, buy_price), ...]}."""
    wb = load_workbook(path, data_only=True)
    sd = wb["Stock Data"]
    longs = wb["Longs"]
    out: dict[str, list[tuple]] = {}

    for i in range(0, 200):
        sym = sd.cell(row=8 + i, column=5).value  # Stock Data!E
        if not sym or not isinstance(sym, str) or not sym.strip():
            break
        sym = sym.strip().upper()

        block_start = 6 + 16 * i
        lots: list[tuple] = []
        rung = 0
        for r in range(block_start + 3, block_start + 13):  # 10 candidate lot rows
            shares = longs.cell(row=r, column=5).value  # E
            price = longs.cell(row=r, column=6).value   # F
            d = longs.cell(row=r, column=4).value        # D
            if shares and price and float(shares) > 0 and float(price) > 0:
                rung += 1
                lots.append((rung, _parse_date(d), float(shares), float(price)))
        if lots:
            out.setdefault(sym, []).extend(lots)  # dup symbol -> append (rare)

    wb.close()
    return out


def read_completed_trades(path: Path):
    wb = load_workbook(path, data_only=True)
    log = wb["Long Log"]
    trades = []
    for r in range(7, 3000):
        b = log.cell(row=r, column=2).value   # B "date ticker"
        if not b:
            break
        shares = log.cell(row=r, column=3).value   # C
        buy = log.cell(row=r, column=4).value      # D
        sell = log.cell(row=r, column=6).value     # F
        completed = log.cell(row=r, column=8).value  # H
        ticker = log.cell(row=r, column=9).value     # I (cached formula)
        if not (shares and buy and sell):
            continue
        sym = (str(ticker).strip().upper() if ticker
               else str(b).split(" ", 1)[-1].strip().upper())
        shares, buy, sell = float(shares), float(buy), float(sell)
        trades.append(
            dict(
                symbol=sym,
                shares=shares,
                buy_price=buy,
                sell_price=sell,
                cost=buy * shares,
                profit=(sell - buy) * shares,
                opened_at=_parse_date(b),
                completed_at=_parse_date(completed) or _parse_date(b),
            )
        )
    wb.close()
    return trades


async def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not path.exists():
        print(f"ERROR: sheet not found at {path}")
        sys.exit(1)

    llc_hash = resolve_llc_hash()
    if not llc_hash:
        print("ERROR: could not resolve the LLC (CASH) account hash — is the token valid?")
        sys.exit(1)
    print(f"Tagging seeded data with LLC account ...{llc_hash[-6:]}")

    await init_db()
    positions = read_positions(path)
    trades = read_completed_trades(path)

    async with SessionLocal() as s:
        # Wipe & reseed the LLC's data (idempotent). Other accounts untouched.
        await s.execute(delete(Lot).where(Lot.account_hash == llc_hash))
        await s.execute(delete(CompletedTrade).where(CompletedTrade.account_hash == llc_hash))
        await s.flush()

        # ticker is global reference data — upsert (don't clobber other accounts' FKs)
        symbols = set(positions) | {t["symbol"] for t in trades}
        for sym in symbols:
            await s.execute(pg_insert(Ticker).values(symbol=sym).on_conflict_do_nothing())

        for sym, lots in positions.items():
            for rung, buy_date, shares, price in lots:
                s.add(Lot(
                    account_hash=llc_hash,
                    symbol=sym,
                    rung=rung,
                    buy_date=buy_date or date.today(),
                    shares=shares,
                    buy_price=price,
                ))
        for t in trades:
            s.add(CompletedTrade(account_hash=llc_hash, **t))

        await s.commit()

    print(f"Seeded {len(positions)} held tickers "
          f"({sum(len(v) for v in positions.values())} lots) and "
          f"{len(trades)} completed trades.")
    for sym, lots in positions.items():
        print(f"  {sym}: {len(lots)} rungs")


if __name__ == "__main__":
    # psycopg async needs the SelectorEventLoop on Windows (see run.py).
    if sys.platform == "win32":
        asyncio.run(main(), loop_factory=asyncio.SelectorEventLoop)
    else:
        asyncio.run(main())
